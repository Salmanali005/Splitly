from fastapi import FastAPI, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, List
from decimal import Decimal
import os
from dotenv import load_dotenv
from datetime import datetime, timedelta
import pytz
import traceback
from fastapi.responses import StreamingResponse
import io
import csv
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter



load_dotenv()

def get_current_time():
    return datetime.now(pytz.UTC)

def now_utc():
    return datetime.now(pytz.UTC)

    

# ============================================
# CREATE APP
# ============================================
app = FastAPI(
    title="Hisaab API",
    description="Trip expense splitting app",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================
# IMPORTS
# ============================================
from database import db
from schemas import *
from services import user as user_service
from services import trip as trip_service
from services import expense as expense_service
from services import debt as debt_service
from services import invitation as invitation_service

# ============================================
# ROOT ENDPOINTS
# ============================================
@app.get("/")
async def root():
    return {"message": "Hisaab API is running"}

@app.get("/health")
async def health_check():
    try:
        db.execute("SELECT 1")
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "database": str(e)}

# ============================================
# DEPENDENCY: Get Current User
# ============================================
def get_current_user(authorization: str = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid token")
    
    token = authorization.split(" ")[1]
    
    try:
        user_id = user_service.get_current_user_id(token)
        user = user_service.get_user_by_id(user_id)
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        return user
    except Exception as e:
        raise HTTPException(status_code=401, detail=str(e))

# ============================================
# AUTH ROUTES
# ============================================
@app.post("/api/auth/register")
async def register(user_data: UserCreate):
    try:
        user = user_service.create_user(
            email=user_data.email,
            name=user_data.name,
            password=user_data.password,
            phone=user_data.phone
        )
        return {"message": "User created successfully", "user": user}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/auth/login")
async def login(login_data: UserLogin):
    try:
        result = user_service.authenticate_user(
            email=login_data.email,
            password=login_data.password
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=401, detail=str(e))

@app.post("/api/auth/refresh")
async def refresh(refresh_token: str):
    try:
        new_token = user_service.refresh_access_token(refresh_token)
        return {"access_token": new_token, "token_type": "bearer"}
    except Exception as e:
        raise HTTPException(status_code=401, detail=str(e))

@app.get("/api/auth/me")
async def get_current_user_route(current_user: dict = Depends(get_current_user)):
    return current_user

@app.put("/api/auth/me")
async def update_profile(user_data: UserUpdate, current_user: dict = Depends(get_current_user)):
    try:
        user = user_service.update_user(current_user['id'], user_data.dict(exclude_unset=True))
        return user
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/auth/change-password")
async def change_password(current_password: str, new_password: str, current_user: dict = Depends(get_current_user)):
    try:
        user_service.change_password(current_user['id'], current_password, new_password)
        return {"message": "Password changed successfully"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================
# TRIP ROUTES
# ============================================
@app.post("/api/trips")
async def create_trip(trip_data: TripCreate, current_user: dict = Depends(get_current_user)):
    try:
        result = trip_service.create_trip(
            user_id=current_user['id'],
            name=trip_data.name,
            destination=trip_data.destination,
            start_date=trip_data.start_date,
            end_date=trip_data.end_date,
            currency=trip_data.currency,
            notes=trip_data.notes
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/trips")
async def get_user_trips(current_user: dict = Depends(get_current_user)):
    return trip_service.get_user_trips(current_user['id'])

@app.get("/api/trips/{trip_id}")
async def get_trip(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    trip = trip_service.get_trip_by_id(trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    return trip

@app.put("/api/trips/{trip_id}")
async def update_trip(trip_id: int, trip_data: TripUpdate, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_trip_admin(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="Only admins can update trips")
    
    result = trip_service.update_trip(trip_id, trip_data.dict(exclude_unset=True))
    if not result:
        raise HTTPException(status_code=404, detail="Trip not found")
    return result

@app.delete("/api/trips/{trip_id}")
async def delete_trip(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_trip_admin(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="Only admins can delete trips")
    
    deleted = trip_service.delete_trip(trip_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Trip not found")
    return {"message": "Trip deleted successfully"}

@app.get("/api/trips/{trip_id}/members")
async def get_trip_members(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return trip_service.get_trip_members(trip_id)

@app.post("/api/trips/{trip_id}/invite")
async def invite_member(trip_id: int, email: str, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_trip_admin(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="Only admins can invite members")
    
    try:
        invitation = trip_service.create_invitation(trip_id, email, current_user['id'])
        return {"message": "Invitation sent", "invitation": invitation}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/trips/{trip_id}/statistics")
async def get_trip_statistics(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return trip_service.get_trip_statistics(trip_id)

# ============================================
# EXPENSE ROUTES
# ============================================
@app.get("/api/categories")
async def get_categories():
    return expense_service.get_expense_categories()

@app.post("/api/trips/{trip_id}/expenses")
async def create_expense(trip_id: int, expense_data: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    try:
        result = expense_service.create_expense(
            trip_id=trip_id,
            paid_by=expense_data.paid_by,
            description=expense_data.description,
            amount=expense_data.amount,
            category=expense_data.category,
            split_type=expense_data.split_type,
            splits=[s.dict() for s in expense_data.splits],
            date=expense_data.date,
            notes=expense_data.notes
        )
        return result
    except Exception as e:
        traceback.print_exc()  # ADD THIS
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/trips/{trip_id}/expenses")
async def get_trip_expenses(trip_id: int, category: Optional[str] = None, member_id: Optional[int] = None, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return expense_service.get_trip_expenses(trip_id, category, member_id)

@app.get("/api/expenses/{expense_id}")
async def get_expense(expense_id: int, current_user: dict = Depends(get_current_user)):
    expense = expense_service.get_expense_by_id(expense_id)
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if not trip_service.is_user_in_trip(current_user['id'], expense['trip_id']):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return expense

@app.put("/api/expenses/{expense_id}")
async def update_expense(expense_id: int, expense_data: ExpenseUpdate, current_user: dict = Depends(get_current_user)):
    expense = expense_service.get_expense_by_id(expense_id)
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if not trip_service.is_user_in_trip(current_user['id'], expense['trip_id']):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    try:
        result = expense_service.update_expense(expense_id, expense_data.dict(exclude_unset=True))
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/api/expenses/{expense_id}")
async def delete_expense(expense_id: int, current_user: dict = Depends(get_current_user)):
    expense = expense_service.get_expense_by_id(expense_id)
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if not trip_service.is_user_in_trip(current_user['id'], expense['trip_id']):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    deleted = expense_service.delete_expense(expense_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}

@app.post("/api/expenses/{expense_id}/comments")
async def add_comment(expense_id: int, text: str, current_user: dict = Depends(get_current_user)):
    expense = expense_service.get_expense_by_id(expense_id)
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if not trip_service.is_user_in_trip(current_user['id'], expense['trip_id']):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    member = db.execute_one(
        "SELECT id FROM members WHERE trip_id = %s AND user_id = %s",
        (expense['trip_id'], current_user['id'])
    )
    
    if not member:
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    comment = expense_service.add_comment_to_expense(expense_id, current_user['id'], member['id'], text)
    return comment

@app.get("/api/expenses/{expense_id}/comments")
async def get_expense_comments(expense_id: int, current_user: dict = Depends(get_current_user)):
    expense = expense_service.get_expense_by_id(expense_id)
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    if not trip_service.is_user_in_trip(current_user['id'], expense['trip_id']):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return expense_service.get_expense_comments(expense_id)

@app.get("/api/trips/{trip_id}/expenses/summary/{member_id}")
async def get_member_expense_summary(trip_id: int, member_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return expense_service.get_member_expense_summary(trip_id, member_id)

# ============================================
# BALANCE & SETTLEMENT ROUTES
# ============================================
@app.get("/api/trips/{trip_id}/balances")
async def get_trip_balances(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return debt_service.calculate_trip_balance_summary(trip_id)

@app.get("/api/trips/{trip_id}/debts/simplified")
async def get_simplified_debts(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return debt_service.get_simplified_debts(trip_id)

@app.get("/api/trips/{trip_id}/debts/my")
async def get_my_debts(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    member = db.execute_one(
        "SELECT id FROM members WHERE trip_id = %s AND user_id = %s",
        (trip_id, current_user['id'])
    )
    
    if not member:
        raise HTTPException(status_code=404, detail="Member not found")
    
    return debt_service.get_owes_summary(trip_id, member['id'])

@app.post("/api/trips/{trip_id}/settlements")
async def create_settlement(trip_id: int, settlement_data: dict, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_trip_admin(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="Only admins can create settlements")
    
    try:
        result = debt_service.create_settlement(
            trip_id=trip_id,
            from_member_id=settlement_data['from_member_id'],
            to_member_id=settlement_data['to_member_id'],
            amount=Decimal(str(settlement_data['amount'])),
            notes=settlement_data.get('notes')
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

        
@app.put("/api/settlements/{settlement_id}/paid")
async def mark_settlement_paid(settlement_id: int, current_user: dict = Depends(get_current_user)):
    try:
        result = debt_service.mark_settlement_paid(settlement_id)
        if not result:
            raise HTTPException(status_code=404, detail="Settlement not found")
        return result
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/trips/{trip_id}/settlements")
async def get_trip_settlements(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return debt_service.get_settlement_history(trip_id)

@app.post("/api/trips/{trip_id}/settlements/auto")
async def auto_settle_debts(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_trip_admin(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="Only admins can settle debts")
    
    settlements = debt_service.settle_all_debts(trip_id)
    return {
        "message": f"Created {len(settlements)} settlements",
        "settlements": settlements
    }

@app.get("/api/trips/{trip_id}/settlements/statistics")
async def get_settlement_statistics(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return debt_service.get_settlement_statistics(trip_id)
# ============================================
# INVITATION ROUTES
# ============================================
@app.get("/api/trips/{trip_id}/invitations")
async def get_trip_invitations(trip_id: int, current_user: dict = Depends(get_current_user)):
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    return trip_service.get_trip_invitations(trip_id)

@app.delete("/api/invitations/{invitation_id}")
async def cancel_invitation(invitation_id: int, current_user: dict = Depends(get_current_user)):
    try:
        # Check if invitation exists and is pending
        invitation = db.execute_one(
            """SELECT i.*, t.id as trip_id
               FROM invitations i
               JOIN trips t ON i.trip_id = t.id
               WHERE i.id = %s AND i.status = 'pending'""",
            (invitation_id,)
        )
        
        if not invitation:
            raise HTTPException(status_code=404, detail="Invitation not found or already processed")
        
        # Check if user is admin of the trip OR the one who sent the invitation
        is_admin = db.execute_one(
            "SELECT id FROM members WHERE trip_id = %s AND user_id = %s AND role = 'admin'",
            (invitation['trip_id'], current_user['id'])
        )
        
        # Use inviter_id (not created_by) - this was the bug!
        if not is_admin and invitation['inviter_id'] != current_user['id']:
            raise HTTPException(status_code=403, detail="You don't have permission to cancel this invitation")
        
        # Update invitation status
        result = db.execute(
            "UPDATE invitations SET status = 'cancelled' WHERE id = %s AND status = 'pending' RETURNING id",
            (invitation_id,)
        )
        
        if not result or len(result) == 0:
            raise HTTPException(status_code=400, detail="Invitation cannot be cancelled")
        
        return {"message": "Invitation cancelled successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/invitations")
async def get_my_invitations(current_user: dict = Depends(get_current_user)):
    """Get all pending invitations for the current user"""
    try:
        user = db.execute_one(
            "SELECT email FROM users WHERE id = %s",
            (current_user['id'],)
        )
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        result = db.execute(
            """SELECT i.*, 
                      t.name as trip_name, 
                      t.destination,
                      u.name as inviter_name
               FROM invitations i
               JOIN trips t ON i.trip_id = t.id
               JOIN users u ON i.inviter_id = u.id
               WHERE i.email = %s
               AND i.status = 'pending'
               AND i.expires_at > NOW()
               ORDER BY i.created_at DESC""",
            (user['email'],)
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/invitations/{token}/accept")
async def accept_invitation_route(token: str, current_user: dict = Depends(get_current_user)):
    """Accept an invitation and join the trip"""
    try:
        # Get user
        user = db.execute_one(
            "SELECT id, email FROM users WHERE id = %s",
            (current_user['id'],)
        )
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Get invitation
        invitation = db.execute_one(
            "SELECT * FROM invitations WHERE token = %s AND status = 'pending'",
            (token,)
        )
        
        if not invitation:
            raise HTTPException(status_code=404, detail="Invalid invitation")
        
        if invitation['email'] != user['email']:
            raise HTTPException(status_code=403, detail="This invitation is for a different email")
        
        if invitation['expires_at'] < now_utc():
            raise HTTPException(status_code=400, detail="Invitation expired")
        
        # Check if already a member
        existing = db.execute_one(
            "SELECT id FROM members WHERE trip_id = %s AND user_id = %s AND is_active = true",
            (invitation['trip_id'], user['id'])
        )
        
        if existing:
            raise HTTPException(status_code=400, detail="You are already a member of this trip")
        
        # Add user as member
        result = db.execute(
            """INSERT INTO members (trip_id, user_id, role, joined_at)
               VALUES (%s, %s, 'viewer', %s)
               RETURNING id""",
            (invitation['trip_id'], user['id'], now_utc())
        )
        
        if not result or len(result) == 0:
            raise HTTPException(status_code=500, detail="Failed to add member")
        
        # Update invitation status
        db.execute(
            "UPDATE invitations SET status = 'accepted' WHERE token = %s",
            (token,)
        )
        
        # Get trip info
        trip = db.execute_one(
            "SELECT id, name, destination FROM trips WHERE id = %s",
            (invitation['trip_id'],)
        )
        
        return {
            "message": "Invitation accepted successfully!", 
            "trip": trip,
            "member_added": True
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/invitations/{token}/decline")
async def decline_invitation_route(token: str, current_user: dict = Depends(get_current_user)):
    """Decline an invitation"""
    try:
        user = db.execute_one(
            "SELECT email FROM users WHERE id = %s",
            (current_user['id'],)
        )
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        invitation = db.execute_one(
            "SELECT * FROM invitations WHERE token = %s AND status = 'pending'",
            (token,)
        )
        
        if not invitation:
            raise HTTPException(status_code=404, detail="Invalid invitation")
        
        if invitation['email'] != user['email']:
            raise HTTPException(status_code=403, detail="This invitation is for a different email")
        
        db.execute(
            "UPDATE invitations SET status = 'cancelled' WHERE token = %s",
            (token,)
        )
        
        return {"message": "Invitation declined"}
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




# ============================================
# EXPORT ROUTES
# ============================================

@app.get("/api/trips/{trip_id}/export/excel")
async def export_trip_excel(trip_id: int, current_user: dict = Depends(get_current_user)):
    """Export trip data as Excel with proper tables"""
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    trip = trip_service.get_trip_by_id(trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    expenses = expense_service.get_trip_expenses(trip_id)
    balances = trip_service.get_trip_balances(trip_id)
    simplified_debts = debt_service.get_simplified_debts(trip_id)
    
    # Create workbook
    wb = Workbook()
    
    # ============ SHEET 1: EXPENSES ============
    ws1 = wb.active
    ws1.title = "Expenses"
    
    # Title
    ws1.merge_cells('A1:F1')
    title_cell = ws1['A1']
    title_cell.value = f"TRIP REPORT: {trip['name'].upper()}"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    ws1['A2'] = f"Destination: {trip['destination'] or 'N/A'}"
    ws1['A2'].font = Font(size=11)
    ws1['A3'] = f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1['A3'].font = Font(size=11)
    
    # Headers
    headers = ["#", "Date", "Description", "Amount", "Paid By", "Category"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    total_amount = 0
    for idx, exp in enumerate(expenses, 1):
        row = 5 + idx
        amount = float(exp['amount'])
        total_amount += amount
        ws1.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal='center')
        ws1.cell(row=row, column=2, value=exp.get('date', '').strftime('%d/%m/%Y') if exp.get('date') else '')
        ws1.cell(row=row, column=3, value=exp['description'])
        ws1.cell(row=row, column=4, value=amount).number_format = '#,##0.00'
        ws1.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        ws1.cell(row=row, column=5, value=exp.get('payer_name', 'Unknown'))
        ws1.cell(row=row, column=6, value=exp.get('category', 'other'))
    
    # Total row
    total_row = 5 + len(expenses) + 1
    ws1.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=total_row, column=4, value=total_amount).font = Font(bold=True)
    ws1.cell(row=total_row, column=4).number_format = '#,##0.00'
    ws1.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(5, total_row + 1):
        for col in range(1, 7):
            ws1.cell(row=row, column=col).border = thin_border
            ws1.cell(row=row, column=col).alignment = Alignment(vertical='center')
    
    # Column widths
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 15
    
    # ============ SHEET 2: BALANCES ============
    ws2 = wb.create_sheet("Balances")
    
    ws2['A1'] = f"BALANCES - {trip['name'].upper()}"
    ws2['A1'].font = Font(size=14, bold=True)
    
    # Headers
    ws2.cell(row=3, column=1, value="Member").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=3, column=2, value="Balance").font = Font(bold=True, color="FFFFFF")
    ws2.cell(row=3, column=1).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws2.cell(row=3, column=2).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    ws2.cell(row=3, column=1).alignment = Alignment(horizontal='center')
    ws2.cell(row=3, column=2).alignment = Alignment(horizontal='center')
    
    for idx, b in enumerate(balances, 1):
        row = 3 + idx
        ws2.cell(row=row, column=1, value=b['name'])
        ws2.cell(row=row, column=2, value=b['balance']).number_format = '#,##0.00'
        if b['balance'] > 0:
            ws2.cell(row=row, column=2).font = Font(color="008000")
        elif b['balance'] < 0:
            ws2.cell(row=row, column=2).font = Font(color="FF0000")
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal='right')
    
    for row in range(3, 3 + len(balances) + 1):
        for col in range(1, 3):
            ws2.cell(row=row, column=col).border = thin_border
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20
    
    # ============ SHEET 3: DEBTS ============
    ws3 = wb.create_sheet("Debts")
    
    ws3['A1'] = f"SIMPLIFIED DEBTS - {trip['name'].upper()}"
    ws3['A1'].font = Font(size=14, bold=True)
    
    if simplified_debts:
        ws3.cell(row=3, column=1, value="From").font = Font(bold=True, color="FFFFFF")
        ws3.cell(row=3, column=2, value="To").font = Font(bold=True, color="FFFFFF")
        ws3.cell(row=3, column=3, value="Amount").font = Font(bold=True, color="FFFFFF")
        ws3.cell(row=3, column=1).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws3.cell(row=3, column=2).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws3.cell(row=3, column=3).fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        ws3.cell(row=3, column=1).alignment = Alignment(horizontal='center')
        ws3.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        ws3.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        
        for idx, debt in enumerate(simplified_debts, 1):
            row = 3 + idx
            ws3.cell(row=row, column=1, value=debt['from_name'])
            ws3.cell(row=row, column=2, value=debt['to_name'])
            ws3.cell(row=row, column=3, value=debt['amount']).number_format = '#,##0.00'
            ws3.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        
        for row in range(3, 3 + len(simplified_debts) + 1):
            for col in range(1, 4):
                ws3.cell(row=row, column=col).border = thin_border
    else:
        ws3['A3'] = "✅ All settled! No debts pending."
        ws3['A3'].font = Font(size=12, bold=True, color="008000")
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 20
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trip_{trip_id}_report.xlsx"}
    )


@app.get("/api/trips/{trip_id}/export/pdf")
async def export_trip_pdf(trip_id: int, current_user: dict = Depends(get_current_user)):
    """Export trip data as PDF with clean formatting"""
    if not trip_service.is_user_in_trip(current_user['id'], trip_id):
        raise HTTPException(status_code=403, detail="You are not a member of this trip")
    
    trip = trip_service.get_trip_by_id(trip_id)
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    expenses = expense_service.get_trip_expenses(trip_id)
    balances = trip_service.get_trip_balances(trip_id)
    simplified_debts = debt_service.get_simplified_debts(trip_id)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=22, textColor=colors.black, alignment=TA_CENTER, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=12, textColor=colors.gray, alignment=TA_CENTER, spaceAfter=20)
    
    elements.append(Paragraph(f"<b>TRIP REPORT</b>", title_style))
    elements.append(Paragraph(f"{trip['name']} • {trip['destination'] or ''}", subtitle_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 15))
    
    # Expenses Table
    elements.append(Paragraph("<b>EXPENSES</b>", styles['Heading2']))
    
    expense_data = [["#", "Date", "Description", "Amount", "Paid By", "Category"]]
    total_amount = 0
    for idx, exp in enumerate(expenses, 1):
        amount = float(exp['amount'])
        total_amount += amount
        expense_data.append([
            str(idx),
            exp.get('date', '').strftime('%d/%m/%Y') if exp.get('date') else '',
            exp['description'],
            f"{amount:,.2f}",
            exp.get('payer_name', 'Unknown'),
            exp.get('category', 'other')
        ])
    
    # Add total row
    expense_data.append(["", "", "TOTAL", f"{total_amount:,.2f}", "", ""])
    
    expense_table = Table(expense_data, colWidths=[0.3*inch, 0.7*inch, 2.5*inch, 0.9*inch, 1.2*inch, 0.8*inch])
    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -2), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.gray),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, -1), (-1, -1), 0.5, colors.gray),
    ]))
    elements.append(expense_table)
    elements.append(Spacer(1, 15))
    
    # Balances Table
    elements.append(Paragraph("<b>MEMBER BALANCES</b>", styles['Heading2']))
    balance_data = [["Member", "Balance"]]
    for b in balances:
        balance_data.append([b['name'], f"{b['balance']:,.2f}"])
    
    balance_table = Table(balance_data, colWidths=[3*inch, 2*inch])
    balance_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.black),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(balance_table)
    elements.append(Spacer(1, 15))
    
    # Simplified Debts Table
    elements.append(Paragraph("<b>SIMPLIFIED DEBTS</b>", styles['Heading2']))
    if simplified_debts:
        debt_data = [["From", "To", "Amount"]]
        for debt in simplified_debts:
            debt_data.append([debt['from_name'], debt['to_name'], f"{debt['amount']:,.2f}"])
        
        debt_table = Table(debt_data, colWidths=[2.2*inch, 2.2*inch, 1.6*inch])
        debt_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(debt_table)
    else:
        elements.append(Paragraph("✅ All settled! No debts pending.", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=trip_{trip_id}_report.pdf"}
    )
# ============================================
# DEBUG ROUTE
# ============================================
@app.get("/api/debug/trip/{trip_id}/members")
async def debug_trip_members(trip_id: int):
    """Debug - check members of a trip"""
    members = db.execute(
        """SELECT m.*, u.name, u.email 
           FROM members m
           JOIN users u ON m.user_id = u.id
           WHERE m.trip_id = %s""",
        (trip_id,)
    )
    return {
        "trip_id": trip_id,
        "member_count": len(members),
        "members": members
    }

@app.get("/api/debug/trip/{trip_id}/balances")
async def debug_trip_balances(trip_id: int):
    """Debug - check balances of a trip"""
    members = db.execute(
        """SELECT m.id, m.user_id, u.name, u.email
           FROM members m
           JOIN users u ON m.user_id = u.id
           WHERE m.trip_id = %s""",
        (trip_id,)
    )
    
    balances = []
    for member in members:
        paid = db.execute_one(
            "SELECT COALESCE(SUM(amount), 0) as total FROM expenses WHERE trip_id = %s AND paid_by = %s",
            (trip_id, member['user_id'])
        )
        owes = db.execute_one(
            """SELECT COALESCE(SUM(es.share_amount), 0) as total
               FROM expense_splits es
               JOIN expenses e ON es.expense_id = e.id
               WHERE e.trip_id = %s AND es.member_id = %s""",
            (trip_id, member['id'])
        )
        
        balances.append({
            'member': member,
            'paid': paid['total'] if paid else 0,
            'owes': owes['total'] if owes else 0,
            'balance': (paid['total'] if paid else 0) - (owes['total'] if owes else 0)
        })
    
    return {
        'trip_id': trip_id,
        'members': members,
        'balances': balances
    }

@app.get("/api/debug/trip/{trip_id}/simplified-debts")
async def debug_simplified_debts(trip_id: int, current_user: dict = Depends(get_current_user)):
    """Debug - check simplified debts"""
    from services import debt as debt_service
    result = debt_service.get_simplified_debts(trip_id)
    return {
        'trip_id': trip_id,
        'balances': trip_service.get_trip_balances(trip_id),
        'simplified_debts': result
    }

# ============================================
# RUN SERVER
# ============================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)